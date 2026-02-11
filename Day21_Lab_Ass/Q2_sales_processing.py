import pandas as pd
from openpyxl import load_workbook

# 1️ Using Pandas

print("Processing using Pandas...")

# Read Excel file
df = pd.read_excel("sales_data.xlsx", sheet_name="2025")

df["Total"] = df["Quantity"] * df["Price"]

df.to_excel("sales_summary_pandas.xlsx", index=False)

print("sales_summary_pandas.xlsx created successfully!")

# 2️ Using OpenPyXL (Without Pandas)

print("Processing using OpenPyXL...")

wb = load_workbook("sales_data.xlsx")
ws = wb["2025"]

ws["D1"] = "Total"

# Calculate total for each row
for row in range(2, ws.max_row + 1):
    quantity = ws[f"B{row}"].value
    price = ws[f"C{row}"].value
    ws[f"D{row}"] = quantity * price

wb.save("sales_summary_openpyxl.xlsx")

print("sales_summary_openpyxl.xlsx created successfully!")
